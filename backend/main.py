from fastapi import FastAPI, Depends, HTTPException, UploadFile, File
from fastapi.middleware.cors import CORSMiddleware
from fastapi.responses import StreamingResponse, PlainTextResponse
from sqlalchemy.orm import Session
from sqlalchemy import select, func, text
from typing import List, Dict
import io, csv
from openpyxl import load_workbook, Workbook

from .settings import settings
from .db import init_db, SessionLocal
from . import models, schemas
from .auth import hash_password, verify_password, create_access_token, get_current_user, admin_required

app = FastAPI(title="Inventory Pro", version="1.0")

origins = [o.strip() for o in settings.ALLOWED_ORIGINS.split(",") if o.strip()]
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins if origins else ["*"],
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)

def get_db():
    db = SessionLocal()
    try: yield db
    finally: db.close()

@app.on_event("startup")
def on_startup():
    init_db()

# ---------- Auth ----------
@app.post("/auth/register", response_model=schemas.UserOut)
def register(user: schemas.UserCreate, db: Session = Depends(get_db)):
    count = db.scalar(select(func.count()).select_from(models.User))
    if count > 0:
        raise HTTPException(status_code=403, detail="Registro cerrado. Un admin debe crear tu usuario.")
    if db.scalar(select(models.User).where(models.User.email == user.email)):
        raise HTTPException(status_code=400, detail="Email ya existe")
    obj = models.User(name=user.name, email=user.email, hashed_password=hash_password(user.password), is_admin=True)
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

@app.post("/auth/login", response_model=schemas.Token)
def login(form: schemas.UserCreate, db: Session = Depends(get_db)):
    u = db.scalar(select(models.User).where(models.User.email == form.email))
    if not u or not verify_password(form.password, u.hashed_password):
        raise HTTPException(status_code=401, detail="Credenciales invÃ¡lidas")
    token = create_access_token({"sub": str(u.id)})
    return {"access_token": token}

@app.get("/me", response_model=schemas.UserOut)
def me(user=Depends(get_current_user)):
    return user

# ---------- Admin usuarios ----------
@app.post("/api/admin/users", response_model=schemas.UserOut)
def admin_create_user(data: schemas.UserAdminCreate, db: Session = Depends(get_db), user=Depends(admin_required)):
    if db.scalar(select(models.User).where(models.User.email == data.email)):
        raise HTTPException(400, "Email ya existe")
    obj = models.User(
        name=data.name, email=data.email,
        hashed_password=hash_password(data.password),
        is_admin=data.is_admin
    )
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

@app.post("/api/admin/users/{user_id}/reset_password", response_model=schemas.UserOut)
def admin_reset_password(user_id: int, data: schemas.ResetPasswordIn, db: Session = Depends(get_db), user=Depends(admin_required)):
    u = db.get(models.User, user_id)
    if not u: raise HTTPException(404, "Usuario no existe")
    u.hashed_password = hash_password(data.password)
    db.commit(); db.refresh(u)
    return u

# ---------- CatÃ¡logos ----------
@app.get("/api/categories", response_model=List[schemas.CategoryOut])
def list_categories(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return db.scalars(select(models.Category)).all()

@app.post("/api/categories", response_model=schemas.CategoryOut)
def create_category(data: schemas.CategoryIn, db: Session = Depends(get_db), user=Depends(admin_required)):
    obj = models.Category(name=data.name)
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

@app.get("/api/warehouses", response_model=List[schemas.WarehouseOut])
def list_warehouses(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return db.scalars(select(models.Warehouse)).all()

@app.post("/api/warehouses", response_model=schemas.WarehouseOut)
def create_warehouse(data: schemas.WarehouseIn, db: Session = Depends(get_db), user=Depends(admin_required)):
    obj = models.Warehouse(name=data.name)
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

@app.get("/api/products", response_model=List[schemas.ProductOut])
def list_products(db: Session = Depends(get_db), user=Depends(get_current_user)):
    return db.scalars(select(models.Product)).all()

@app.post("/api/products", response_model=schemas.ProductOut)
def create_product(data: schemas.ProductIn, db: Session = Depends(get_db), user=Depends(admin_required)):
    obj = models.Product(**data.model_dump())
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

# ---------- Movimientos & Stock ----------
def apply_stock(db: Session, product_id: int, warehouse_id: int, delta: int):
    s = db.get(models.Stock, {"product_id": product_id, "warehouse_id": warehouse_id})
    if s: s.qty = (s.qty or 0) + delta
    else:
        s = models.Stock(product_id=product_id, warehouse_id=warehouse_id, qty=delta)
        db.add(s)

@app.post("/api/movements", response_model=schemas.MovementOut)
def create_movement(data: schemas.MovementIn, db: Session = Depends(get_db), user=Depends(get_current_user)):
    if data.type not in ("in", "out"):
        raise HTTPException(400, "Tipo invÃ¡lido")
    delta = data.qty if data.type == "in" else -abs(data.qty)
    apply_stock(db, data.product_id, data.warehouse_id, delta)
    obj = models.Movement(**data.model_dump())
    db.add(obj); db.commit(); db.refresh(obj)
    return obj

# ---------- Reportes ----------
@app.get("/api/reports/stock")
def stock_report(db: Session = Depends(get_db), user=Depends(get_current_user)):
    stmt = text('''
        SELECT p.id, p.code, p.name, COALESCE(SUM(s.qty),0) AS total
        FROM products p
        LEFT JOIN stocks s ON s.product_id = p.id
        GROUP BY p.id, p.code, p.name
        ORDER BY p.code''')
    rows = db.execute(stmt).mappings().all()
    return [dict(r) for r in rows]

@app.get("/api/reports/stock_matrix")
def stock_matrix(db: Session = Depends(get_db), user=Depends(get_current_user)):
    stmt = text('''
        SELECT p.id AS product_id, p.code, p.name, w.name AS warehouse, COALESCE(s.qty,0) as qty
        FROM products p
        CROSS JOIN warehouses w
        LEFT JOIN stocks s ON s.product_id = p.id AND s.warehouse_id = w.id
        ORDER BY p.code, w.name''')
    data = db.execute(stmt).mappings().all()
    matrix = {}
    for r in data:
        key = (r["product_id"], r["code"], r["name"])
        if key not in matrix:
            matrix[key] = {"product_id": r["product_id"], "code": r["code"], "name": r["name"], "by_warehouse": {}, "total": 0}
        matrix[key]["by_warehouse"][r["warehouse"]] = r["qty"]
        matrix[key]["total"] += r["qty"]
    return list(matrix.values())

# ---------- Export CSV/Excel ----------
@app.get("/api/export/stock.csv")
def export_stock_csv(db: Session = Depends(get_db), user=Depends(get_current_user)):
    data = stock_matrix(db, user)  # reuse
    # collect warehouse headers
    wh = set()
    for row in data:
        for k in row["by_warehouse"].keys():
            wh.add(k)
    headers = ["product_id","code","name",*sorted(wh), "total"]
    sio = io.StringIO()
    writer = csv.writer(sio)
    writer.writerow(headers)
    for row in data:
        line = [row["product_id"], row["code"], row["name"]]
        for h in sorted(wh):
            line.append(row["by_warehouse"].get(h,0))
        line.append(row["total"])
        writer.writerow(line)
    return PlainTextResponse(sio.getvalue(), media_type="text/csv")

@app.get("/api/export/stock.xlsx")
def export_stock_xlsx(db: Session = Depends(get_db), user=Depends(get_current_user)):
    data = stock_matrix(db, user)
    # warehouses
    wh = set()
    for row in data:
        for k in row["by_warehouse"].keys():
            wh.add(k)
    headers = ["product_id","code","name",*sorted(wh), "total"]
    wb = Workbook()
    ws = wb.active
    ws.title = "Stock"
    ws.append(headers)
    for row in data:
        line = [row["product_id"], row["code"], row["name"]]
        for h in sorted(wh):
            line.append(row["by_warehouse"].get(h,0))
        line.append(row["total"])
        ws.append(line)
    bio = io.BytesIO()
    wb.save(bio); bio.seek(0)
    return StreamingResponse(bio, media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition":"attachment; filename=stock.xlsx"})

# ---------- Importar existencias desde Excel ----------
# Formato esperado (hoja 1): headers: code, warehouse, qty
@app.post("/api/import/stock")
def import_stock(file: UploadFile = File(...), db: Session = Depends(get_db), user=Depends(admin_required)):
    content = file.file.read()
    wb = load_workbook(io.BytesIO(content))
    ws = wb.active
    headers = [str(c.value).strip().lower() for c in next(ws.iter_rows(min_row=1, max_row=1))[0:3]]
    if headers != ["code","warehouse","qty"]:
        raise HTTPException(400, "Encabezados esperados: code, warehouse, qty")
    for row in ws.iter_rows(min_row=2):
        code = str(row[0].value).strip() if row[0].value is not None else None
        wh_name = str(row[1].value).strip() if row[1].value is not None else None
        try:
            qty = int(row[2].value)
        except:
            qty = 0
        if not code or not wh_name: 
            continue
        # producto
        prod = db.scalar(select(models.Product).where(models.Product.code==code))
        if not prod:
            prod = models.Product(code=code, name=code)
            db.add(prod); db.flush()
        # bodega
        wh = db.scalar(select(models.Warehouse).where(models.Warehouse.name==wh_name))
        if not wh:
            wh = models.Warehouse(name=wh_name)
            db.add(wh); db.flush()
        # set absolute qty
        st = db.get(models.Stock, {"product_id": prod.id, "warehouse_id": wh.id})
        if not st:
            st = models.Stock(product_id=prod.id, warehouse_id=wh.id, qty=qty)
            db.add(st)
        else:
            st.qty = qty
    db.commit()
    return {"status":"ok","message":"Stock actualizado"}

